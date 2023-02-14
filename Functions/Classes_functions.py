# Main Classes, functions, APIs

import mysql.connector
# import zillow
import datetime
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
import locale
import openpyxl
import pprint
import logging
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver

street = '501 clear blue way'
city = 'mcdonough'
short_state = 'GA'
state = 'Georgia'
randomid = 'alex2'
xls_name = 'HomeAnalyzer.xlsx'

class HouseData(object):
    def __init__(self, street, state, city, short_state, xls_name, randomid):
        self.driver = webdriver.Chrome('/Users/alexdezho/Documents/Personal/chromedriver')
        self.street = street
        self.state = state
        self.city = city
        self.xls_name = xls_name
        self.short_state = short_state
        self.randomid = randomid
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.replace(" ", "").lower() + " " + self.short_state.lower()
        self.googleMaps_url = "https://www.google.com/maps/"
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'state': self.state,
            'coordinates': '',
            'id': self.randomid
        }

    def printparams(self):
        print(self.dict_basic_info)

    # can be switched with Google API for Coordinates
    def google_maps_addr_coord(self):
        try:
            # opening Google Maps and zooming into the address to get the correct coordinates
            driver = self.driver
            driver.get(self.googleMaps_url)
            driver.maximize_window()
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="searchboxinput"]')))
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(self.full_addr)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="searchbox-searchbutton"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom1
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom2
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom3
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom4
            time.sleep(3)

            # getting the coordinates from url string
            self.coordurl = self.driver.current_url
            index = self.coordurl.find("@") + 1
            self.coordurl = self.coordurl[index:index + 21]
            self.coordurl = self.coordurl.replace(",", ":")
            self.dict_basic_info['coordinates'] = self.coordurl
            print("google coordinates is: {}".format(self.dict_basic_info['coordinates']))
        except:
            print('failed to locate google coordinates')
            return False

    def closebrowser(self):
        self.driver.close()
        print('Browser closed')

    # creating a new xls sheet for the address
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        print(wb.sheetnames)
        if wb.sheetnames.count(self.full_addr[:25]) == 0:
            example_sheet = wb["Sheet1 Copy Copy Copy"]
            wb.copy_worksheet(example_sheet).title = self.full_addr[:25]
            wb.save(self.xls_name)
            print("xls sheet is ready: {}".format(self.full_addr[:25]))
            print(wb.sheetnames)
            wb.close()
            return True
        else:
            print("address is exists in xls data")
            wb.close()
            return False

    # copy all dictionaries to xls file
    def basic_info_dict_to_xls(self):
        # opening xls
        print('opening xls: {}'.format(self.xls_name))
        wb = openpyxl.load_workbook(self.xls_name)
        print(wb.sheetnames)
        sheet = wb[self.full_addr[:25]]
        sheet['B2'].value = self.dict_basic_info['street']
        sheet['B3'].value = self.dict_basic_info['city']
        sheet['B4'].value = self.dict_basic_info['state']
        sheet['B5'].value = self.dict_basic_info['coordinates']
        sheet['B6'].value = self.dict_basic_info['id']
        wb.save(self.xls_name)
        wb.close()
        return True

    # connecting to mysql and copy data into a table
    def mysql(self):
        print('Connect to MySQL server')
        db = mysql.connector.connect(
            host='localhost',  # ip when it will be on cloud
            user='root',
            passwd='NV27vnmc',
            database='data_list_storage'
        )
        print(db)  # checking our connection to DB
        mycursor = db.cursor()
        sql = "INSERT INTO Houses (street, city, state, coordinates, id) VALUES (%s, %s, %s, %s, %s)"
        val = (self.dict_basic_info['street'],
               self.dict_basic_info['city'],
               self.dict_basic_info['state'],
               self.dict_basic_info['coordinates'],
               self.dict_basic_info['id'])
        mycursor.execute(sql, val)
        db.commit()
        print('address inserted to mysQL')
        print(mycursor.statement)


def automation_tool(street, state, city, short_state, xls_name, randomid):
    hd = HouseData(street, state, city, short_state, xls_name, randomid)
    hd.google_maps_addr_coord()
    hd.printparams()
    hd.xls_new_sheet_for_search_create()
    hd.basic_info_dict_to_xls()
    hd.mysql()


automation_tool(street, state, city, short_state, xls_name, randomid)










class Crime(object):
    def __init__(self, street, state, city, short_state, xls_name):
        # all setup params
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.onboardnavigator_url = 'http://www.onboardnavigator.com/webcontent/OBWC_Search.aspx?&AID=102'
        self.city_data_url = 'http://www.city-data.com'
        self.home_facts_url = 'https://www.homefacts.com/'
        self.neighborhoodscout_url = 'https://www.neighborhoodscout.com/' + self.short_state.lower() + '/' + self.city.lower() + '/crime'
        self.bestplaces_url = 'https://www.bestplaces.net/crime/city/' + self.state.lower() + '/' + self.city.lower()
        # add NA
        #dictionaries
        self.dict_crime_total = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'zip_code': 'NA',
            'metropolitan': 'NA',
            'link_google_maps': 'NA'
        }
        self.dict_onboardnavigator = {
            'Total personal': 'NA',
            'Total property': 'NA',
            'Total overall': 'NA',
            'Year': '2019',

        }
        self.dict_city_data = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Year': '2019',

        }
        self.dict_home_facts = {
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Year': '2019',

        }
        self.dict_offenders = {
            'offender1': 'NA',
            'offender2': 'NA',
            'offender3': 'NA',

        }
        self.dict_neighborhoodscout = {
            'Diagram': 'NA',
            'List of safe areas': 'NA',
        }
        self.dict_bestplaces = {
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')
    # the functions below written in a working flow
    # getting all the information and copy into dicts
    def onboardnavigator_to_dict(self):
        try:
            print('onboardnavigator')
            driver = self.driver
            driver.get(self.onboardnavigator_url)
            time.sleep(10)
            print('Navigator tool opened')
            # select state
            state = driver.find_element_by_xpath('//*[@id="ddlGenLookupStateID"]').click()
            time.sleep(5)
            Select(driver.find_element_by_tag_name('select')).select_by_visible_text(self.state)
            time.sleep(5)
            print('state selected')
            driver.find_element_by_xpath('//*[@id="tbGenSearch"]').send_keys(self.city)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="radGenCity"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="cmdGenSave"]').click()
            time.sleep(10)
            print('navigator address located')
            link = driver.current_url
            self.dict_onboardnavigator['Total personal'] = link
            self.dict_onboardnavigator['Total property'] = link
            self.dict_onboardnavigator['Total overall'] = link
            print('onboardnavigator params was copied to dictionary , success {}'.format(self.dict_onboardnavigator))
        except:
            print('failed to locate navigator')

    def city_data_to_dict(self):
        try:
            print('citydata')
            driver = self.driver
            driver.get(self.city_data_url)
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').send_keys(self.city + ' ' + self.state)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="search_bar_box"]/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,4100)")
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="sex-offenders"]/p')))
            # select city data elemnets and copy to dictionary
            self.dict_city_data['total info'] = driver.find_element_by_xpath('//*[@id="sex-offenders"]/p').text
            self.dict_city_data['Pic of graph'] = driver.current_url
            self.dict_crime_total['total info'] = self.dict_city_data['total info']
            self.dict_crime_total['Pic of graph'] = self.dict_city_data['Pic of graph']
            print('city_data total info copied {}'.format(self.dict_city_data))
        except:
            print('failed to locate city data elements')
        try:
            driver = self.driver
            self.dict_city_data['Crime Index city'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[15]').text
            self.dict_city_data['US avarage'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[1]').text
            self.dict_crime_total['Crime Index city'] = self.dict_city_data['Crime Index city']
            self.dict_crime_total['US avarage'] = self.dict_city_data['US avarage']
            print('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            logging.debug('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            return True
        except:
            self.dict_city_data['Crime Index city'] = 'Crime table not exists in city_data for this state'
            print('Crime table not exists in city_data for this state')
            logging.debug('fail')
            return False

    def home_facts_to_dict(self):
        try:
            print('homefacts')
            driver = self.driver
            driver.get(self.home_facts_url)
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            addr = driver.find_element_by_xpath('//*[@id="fulladdress"]')
            addr.click()
            time.sleep(3)
            addr.send_keys(self.full_addr)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="main-search-form"]/div/div/div/div[1]/span/button').click()
            time.sleep(10)
            element = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div[2]/div[1]/div[3]/ul/li[1]/span[4]/a')
            driver.execute_script("window.scrollTo(0,600)")
            time.sleep(3)
            element.click()
            print(driver.current_url)
            print('view crime statistics report')
            time.sleep(10)
            try:
                print('trying to click')
                driver.find_element_by_partial_link_text('view crime statistics report').click()
            except:
                print('trying to click with second option')
                driver.execute_script("window.scrollTo(0,2700)")
                time.sleep(7)
                driver.find_element_by_partial_link_text('view crime statistics report').click()

            time.sleep(10)
            print(driver.current_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="crimeScore"]/div[1]/div[4]')))
            self.dict_home_facts['Overall Score'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[4]').get_attribute('class')
            self.dict_home_facts['Overall score big num'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[2]').text
            self.dict_home_facts['Score small procents'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[3]').text
            self.dict_home_facts['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_home_facts['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_home_facts['Score small procents'] = self.dict_home_facts['Score small procents']
            print(self.dict_home_facts['Overall Score'])
            print(self.dict_home_facts['Overall score big num'])
            print(self.dict_home_facts['Score small procents'])
            self.dict_crime_total['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_crime_total['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_crime_total['Score small procents'] = self.dict_home_facts['Score small procents']
            print('dict_home_facts params was copied to dictionary , success {}'.format(self.dict_home_facts))
            print('dict_offenders params was copied to dictionary , success {}'.format(self.dict_offenders))
        except:
            print('failed to locate and copy from home facts')

    def neighborhoodscout_to_dict(self):
        try:
            print('neighborhoodscout')
            print(self.neighborhoodscout_url)
            data = requests.get(self.neighborhoodscout_url)
            time.sleep(5)
            soup = BeautifulSoup(data.content, 'html.parser')
            list = soup.find_all('script', type='application/ld+json')
            list = str(list)
            index_list_start = list.find('itemListOrder')
            new_list = list[index_list_start:]
            index_list_end = new_list.find('</script>')
            orig_list = new_list[:index_list_end]
            index1 = orig_list.find('[')
            orig_list = orig_list[index1:]
            index2 = orig_list.find(']')
            # list of safety places taken from HTML converted to string
            orig_list = orig_list[index1:index2]
            self.dict_neighborhoodscout['List of safe areas'] = orig_list
            self.dict_neighborhoodscout['Diagram'] = self.neighborhoodscout_url
            print('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            logging.debug('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            return True
        except:
            logging.debug('fail to connect or copy from neighborhoodscout')
            print('fail to connect or copy from neighborhoodscout')
            return False

    def bestplaces_to_dict(self):
        try:
            print('bestplaces')
            driver = self.driver
            driver.get(self.bestplaces_url)
            time.sleep(5)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]')))
            self.dict_bestplaces['Violent crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]').text
            self.dict_bestplaces['Property crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[2]').text

            # Photos and Maps
            driver.find_element_by_xpath('//*[@id="form1"]/div[5]/div/div/p[3]/a[3]/u').click()
            time.sleep(4)
            self.dict_bestplaces['Photos and Maps of the city'] = driver.current_url
            print('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))
            logging.debug('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))

            self.dict_crime_total['Violent crime & US average'] = self.dict_bestplaces['Violent crime & US average']
            self.dict_crime_total['Property crime & US average'] = self.dict_bestplaces['Property crime & US average']
            self.dict_crime_total['Photos and Maps of the city'] = self.dict_bestplaces['Photos and Maps of the city']

            return True
        except:
            logging.debug('fail to connect or copy from bestplaces')
            print('fail to connect or copy from bestplaces')
            return False

    # print all dictionaries
    def printall(self):
        print('All dictionaries\n')
        pp = pprint.PrettyPrinter(indent=4)
        print(self.dict_basic_info)
        pp.pprint(self.dict_basic_info)
        print(self.dict_onboardnavigator)
        pp.pprint(self.dict_onboardnavigator)
        print(self.dict_city_data)
        pp.pprint(self.dict_city_data)
        print(self.dict_home_facts)
        pp.pprint(self.dict_home_facts)
        print(self.dict_offenders)
        pp.pprint(self.dict_offenders)
        print(self.dict_neighborhoodscout)
        pp.pprint(self.dict_neighborhoodscout)
        print(self.dict_bestplaces)
        pp.pprint(self.dict_bestplaces)
        return True
    # returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info

    def return_dict_onboardnavigator(self):
        return self.dict_onboardnavigator

    def return_dict_city_data(self):
        return self.dict_city_data

    def return_dict_home_facts(self):
        return self.dict_home_facts

    def return_dict_offenders(self):
        return self.dict_offenders

    def return_dict_neighborhoodscout(self):
        return self.dict_neighborhoodscout

    def return_dict_bestplaces(self):
        return self.dict_bestplaces

    def return_dict_crime_total(self):
        return self.dict_crime_total

    # copy all dictionaries to xls file
    def xls_new_sheet_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.full_addr[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.full_addr[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
            wb.close()
            return True
        else:
            print("address is already exists in database!, recopy new run info ")
            logging.debug("address is already exists in database!, recopy new run info ")
            return False

    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.full_addr[:25]]

            sheet['A3'].value = self.dict_basic_info['street']
            sheet['C3'].value = self.dict_basic_info['city']
            sheet['D3'].value = self.dict_basic_info['state']

            sheet['B24'].value = self.dict_onboardnavigator['Total personal']
            sheet['B25'].value = self.dict_onboardnavigator['Total property']
            sheet['B26'].value = self.dict_onboardnavigator['Total overall']
            sheet['B27'].value = self.dict_onboardnavigator['Year']

            sheet['B29'].value = self.dict_city_data['Crime Index city']
            sheet['B30'].value = self.dict_city_data['US avarage']
            sheet['B31'].value = self.dict_city_data['Pic of graph']
            sheet['B32'].value = self.dict_city_data['total info']

            sheet['B34'].value = self.dict_home_facts['Overall Score']
            sheet['B35'].value = self.dict_home_facts['Overall score big num']
            sheet['B36'].value = self.dict_home_facts['Score small procents']
            sheet['B37'].value = self.dict_offenders['offender1']
            sheet['B38'].value = self.dict_offenders['offender2']
            sheet['B39'].value = self.dict_offenders['offender3']
            sheet['B40'].value = self.dict_home_facts['Year']

            sheet['B42'].value = self.dict_neighborhoodscout['Diagram']
            sheet['B43'].value = self.dict_neighborhoodscout['List of safe areas']

            sheet['B45'].value = self.dict_bestplaces['Violent crime & US average']
            sheet['B46'].value = self.dict_bestplaces['Property crime & US average']
            sheet['B47'].value = self.dict_bestplaces['Photos and Maps of the city']

            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Elements saved in {}".format(self.xls_name))
            logging.debug("Elements saved in {}".format(self.xls_name))
            return True
        except:
            return False
class Schools(object):
    def __init__(self, street, state, city, short_state, xls_name, county_name, zip_code):
        # all setup params
        self.zip_code = zip_code
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.county = county_name
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.greatschools_url = 'https://www.greatschools.org/'
        self.schooldigger_url = 'https://www.schooldigger.com/'
        self.homefacts_url = 'https://www.homefacts.com/'
        self.niche_url = 'https://www.niche.com/?ref=k12'

        self.dict_schools_general = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA',
            'school - HF elementary name': 'NA',
            'school - HF elementary link': 'NA',
            'school - HF middle name': 'NA',
            'school - HF middle link': 'NA',
            'school - HF high name': 'NA',
            'school - HF high link': 'NA'

        }


        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'county': self.county
        }
        self.dict_greatschools = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_schooldigger = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'
        }
        self.dict_homefacts = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_niche = {
            'link - County Schools': 'NA',
            'name - global': 'NA',
            'rank - School Districts if exists': 'NA',
            'grade - overall niche grade': 'NA',
            'link - all ranks state county schools/metropolitan/national': 'NA'

        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def greateschools_to_dict(self):
        try:  # connecting to greateschools
            driver = self.driver
            driver.get(self.greatschools_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input')))
            driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
            driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
            time.sleep(5)
            print(driver.current_url)
            try:
                # elementary school assigned tags
                time.sleep(10)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Elementary School')))
                driver.find_element_by_partial_link_text('Elementary School').click()
                time.sleep(3)
                self.dict_greatschools['school - elementary link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                print(self.dict_greatschools['school - elementary link'])
                school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                self.dict_greatschools['school - elementary name'] = school_name
                print('Elementary school name: {}'.format(school_name))
                self.dict_schools_general['school - elementary link'] = self.dict_greatschools['school - elementary link']
                self.dict_schools_general['school - elementary name'] = self.dict_greatschools['school - elementary name']
                driver.back()
            except:
                print('failed to locate elemantary school from greateschools')

            try:
                # middle school assigned tags
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Middle School')))
                driver.find_element_by_partial_link_text('Middle School').click()
                time.sleep(3)
                self.dict_greatschools['school - middle link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                print(self.dict_greatschools['school - middle link'])
                school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                print('Middle school name: {}'.format(school_name))
                self.dict_greatschools['school - middle name'] = school_name
                self.dict_schools_general['school - middle link'] = self.dict_greatschools['school - middle link']
                self.dict_schools_general['school - middle name'] = self.dict_greatschools['school - middle name']
                driver.back()
            except:
                print('failed to locate middle school from greateschools')
                self.dict_greatschools['school - middle link'] = 'NA'


            try:
                # high school assigned tags
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'High School')))
                driver.find_element_by_partial_link_text('High School').click()
                time.sleep(3)
                self.dict_greatschools['school - high link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                print(self.dict_greatschools['school - high link'])
                school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                print('High school name: {}'.format(school_name))
                self.dict_greatschools['school - high name'] = school_name
                self.dict_schools_general['school - high link'] = self.dict_greatschools['school - high link']
                self.dict_schools_general['school - high name'] = self.dict_greatschools['school - high name']
                driver.back()
            except:
                print('failed to locate high school from greateschools')
        except:
            print('something went wrong with greateschools')

    def schooldigger_to_dict(self): #check
        try:
            driver = self.driver
            driver.get(self.schooldigger_url)
            print(driver.current_url)
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtHPAC"]')))
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(Keys.ENTER)
            time.sleep(3)
            print(driver.current_url)

            # elementary schools under boundary tags
            try:
                driver.find_element_by_partial_link_text('Elementary').click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - elementary link'] = driver.current_url
                self.dict_schooldigger['school - elementary name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                print('elemantary school found in schooldigger')
                driver.back()
                time.sleep(10)
            except:
                print('elemantary school was not fount in schooldigger')

            try:
                driver.find_element_by_partial_link_text('Middle').click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - middle link'] = driver.current_url
                self.dict_schooldigger['school - middle name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                driver.back()
                time.sleep(10)
                print('middle school found in schooldigger')
            except:
                print('middle school was not fount in schooldigger')
            # high boundary tags
            try:
                driver.find_element_by_partial_link_text('High').click()
                time.sleep(10)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - high link'] = driver.current_url
                self.dict_schooldigger['school - high name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                print('high school found in schooldigger')
                driver.back()
                time.sleep(10)
            except:
                print('high school was not fount in schooldigger')
            print('schooldigger params was copied to dictionary , success {}'.format(self.dict_schooldigger))
        except:
            print('failed to connect or locate params from schooldigger')

    def homefacts_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.homefacts_url)
            time.sleep(10)
            print('homefacts entered ')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            driver.find_element_by_xpath('//*[@id="fulladdress"]').click()
            time.sleep(2)
            # driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(str(self.zip_code))
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(Keys.ENTER)
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="navbar"]/ul/li[4]/a')))
            driver.find_element_by_xpath('//*[@id="navbar"]/ul/li[4]/a').click()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,550)")
            time.sleep(3)
            print(driver.current_url)
            time.sleep(3)
            print('schools list located')
            # elementary
            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'ELEMENTARY SCHOOL')))
                driver.find_element_by_partial_link_text('ELEMENTARY SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('ELEMENTARY {}'.format(school_name))
                time.sleep(3)
                self.dict_homefacts['school - elementary name'] = school_name #                 //*[@id="school_year_2019"]/div[1]/div[2]
                self.dict_homefacts['school - elementary link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - elementary link'])
                time.sleep(2)
                self.dict_schools_general['school - HF elementary name'] = self.dict_homefacts['school - elementary name']
                self.dict_schools_general['school - HF elementary link'] = self.dict_homefacts['school - elementary link']
                print('elemantary school found in homefacts')
                driver.back()
                time.sleep(5)
            except:
                print('elemantary school was not fount in homefacts')

            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'MIDDLE SCHOOL')))
                driver.find_element_by_partial_link_text('MIDDLE SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('middle {}'.format(school_name))
                self.dict_homefacts['school - middle name'] = school_name
                self.dict_homefacts['school - middle link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - middle link'])
                time.sleep(2)
                self.dict_schools_general['school - HF middle name'] = self.dict_homefacts['school - middle name']
                self.dict_schools_general['school - HF middle link'] = self.dict_homefacts['school - middle link']
                driver.back()
                time.sleep(5)
            except:
                print('middle school was not fount in homefacts')

            try:
                print('trying to locate high school')
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'HIGH SCHOOL')))
                driver.find_element_by_partial_link_text('HIGH').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('High school {}'.format(school_name))
                time.sleep(5)
                self.dict_homefacts['school - high name'] = school_name
                print('trying to locate high school grade from pic') # //*[@id="school_year_2018"]/div[1]/div[2]
                self.dict_homefacts['school - high link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - high link'])
                time.sleep(2)
                self.dict_schools_general['school - HF high name'] = self.dict_homefacts['school - high name']
                self.dict_schools_general['school - HF high link'] = self.dict_homefacts['school - high link']
                driver.back()
                time.sleep(5)
            except:
                print('high school was not fount in homefacts')

            print('homefacts params was copied to dictionary , success {}'.format(self.dict_homefacts))
        except:
            print('fail to copy params from homefacts')

    def niche_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.niche_url)
            time.sleep(4)
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').click()
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').send_keys(
                self.county + ' county')
            time.sleep(1)
            # driver.find_element_by_xpath('//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').sendKeys(Keys.ENTER)
            time.sleep(4)
            self.dict_niche['link - County Schools'] = driver.current_url
            self.dict_niche['name - global'] = driver.current_url
            self.dict_niche['rank - School Districts if exists'] = driver.current_url
            self.dict_niche['grade - overall niche grade'] = driver.current_url
            self.dict_niche['link - all ranks state county schools/metropolitan/national'] = driver.current_url

            print('niche params was copied to dictionary , success {} '.format(self.dict_niche))
            return True
        except:
            print('fail to locate params from niche')
            logging.debug('fail')
            return False

    def printall(self):
        pp = pprint.PrettyPrinter(indent=4)
        print('greate schools')
        pp.pprint(self.dict_greatschools)
        print('school digger')
        pp.pprint(self.dict_schooldigger)
        print('home facts')
        pp.pprint(self.dict_homefacts)
        print('niche')
        pp.pprint(self.dict_niche)

    # returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info
    def return_dict_greateshcools(self):
        return self.dict_greatschools
    def return_dict_schooldigger(self):
        return self.dict_schooldigger
    def return_dict_homefacts(self):
        return self.dict_homefacts
    def return_dict_niche(self):
        return self.dict_niche
    def return_dict_schools_general(self):
        return self.dict_schools_general
    # copy all dictionaries to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.full_addr[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.full_addr[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet name: {}".format(self.full_addr[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database recopy new params")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        print('copy dicts to xls')
        wb = openpyxl.load_workbook(self.xls_name)
        sheet = wb[self.full_addr[:25]]
        # print(wb.sheetnames)
        sheet['F24'].value = self.dict_greatschools['school - elementary name']
        sheet['F25'].value = self.dict_greatschools['school - elementary link']
        sheet['F26'].value = self.dict_greatschools['school - middle name']
        sheet['F27'].value = self.dict_greatschools['school - middle link']
        sheet['F28'].value = self.dict_greatschools['school - high name']
        sheet['F29'].value = self.dict_greatschools['school - high link']

        sheet['F31'].value = self.dict_schooldigger['school - elementary name']
        sheet['F32'].value = self.dict_schooldigger['school - elementary link']
        sheet['F33'].value = self.dict_schooldigger['school - middle name']
        sheet['F34'].value = self.dict_schooldigger['school - middle link']
        sheet['F35'].value = self.dict_schooldigger['school - high name']
        sheet['F36'].value = self.dict_schooldigger['school - high link']

        sheet['F38'].value = self.dict_homefacts['school - elementary name']
        sheet['F39'].value = self.dict_homefacts['school - elementary link']
        sheet['F40'].value = self.dict_homefacts['school - middle name']
        sheet['F41'].value = self.dict_homefacts['school - middle link']
        sheet['F42'].value = self.dict_homefacts['school - high name']
        sheet['F43'].value = self.dict_homefacts['school - high link']

        # sheet['F29'].value = self.dict_basic_info['street']
        # sheet['F30'].value = self.dict_basic_info['city']
        # sheet['F31'].value = self.dict_basic_info['state']

        wb.save(self.xls_name)
        wb.close()
        # printing the process
        print("Dictionaries was completed & saved in {}".format(self.xls_name))
        logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
        return True
class Builders(object):
    def __init__(self, metropolitan, short_state, xls_name):
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")
        self.lennar_url = 'https://www.lennar.com/'  # builders website
        self.metropolitan = metropolitan.lower() + ' ' + short_state.lower()  # full name for search
        self.floorplan_homes = ''
        self.xls_name = xls_name  # xls name
        self.short_state = short_state
        self.clicked = ''
        self.list_of_homes = []
        self.community_address_list_full = []  # full list of community addresses
        self.community_address_list_names = []  # full list of community names
        self.id_random_list = []
        self.row = 2
        self.rowhome = 2
        self.general_row = 0
        self.row_num_xls = 0
        self.index = 1
        self.x_path_name_to_scroll = ''
        self.element = ''  # scrolling element
        self.addr = ''
        self.name = ''
        self.update_time = ''  # update time
        self.homes_urls = []  # list of homes urls
        self.x_path_name = ''
        self.num_of_communities = ''
        self.num_of_pages = ''  # num of community pages
        self.num_of_comm_pages = ''
        self.num_of_homes_pages = ''  # num of homes pages
        self.num_of_moving_homes = ''  # num of homes
        self.dict_lennar_filter_info = {
            'Communities num': '',
            'metropolitan name': metropolitan,
            'Quick Move-In Homes num': '',
            'Floorplans num': '',
            'time of update': ''
        }

        # community data for mysql and xls
        self.dict_community_data = {
            'address': '',
            'name_community': '',
            'overview': '',
            'approximate_hoa_fees': 'non',
            'approximate_tax_rate': 'non',
            'included_features_pdf_url': 'under solution',
            'community_map_url': 'no pic',
            'community_home_picture_for_present_url': 'no pic',
            'available_homes_quick_move_in_homes': '',
            'available_homes_floorplans': '',
            'id_generated':''

        }

        # home data for mysql and xls
        self.dict_home_data = {
            'address': '',
            'name_community': metropolitan,
            'home_name': '',
            'home_site': '',
            'availability': '',
            'priced_from': '',
            'home_size': '',
            'stories': '',
            'beds': '',
            'type': '',
            'baths': '',
            'garage': '',
            'id': '',
            'id_generated': '',
            'description': '',
            'included_features_pdf': 'under solution',
            'floorplans_with_furniture_pic': '',
            'id_generated_home': '',
            'gallery_view_picture': ''
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def lennar_filter_and_toolbar_info_copy(self):
        driver = self.driver
        driver.get(self.lennar_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input')))  # await command
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').click()
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').send_keys(self.metropolitan)
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/button').send_keys(Keys.ENTER)
        print('Connected to Lennar')
        time.sleep(3)
        # Create filter
        print('creating filter')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]')))
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]').click()
        time.sleep(3)
        # community type
        try:
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/a').click()
            time.sleep(3)
        except:
            print('no community type')
        # add single family loop
        for i in range(0, 10):
            print(i)
            element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div/div/ul/li[' + str(i) + ']/label'
            try:
                filter = driver.find_element_by_xpath(element).text
            except:
                print('not such element exists')
                filter = 'NO'

            if filter == 'Single Family':
                mainelem = element
                filter = driver.find_element_by_xpath(mainelem)
                filter.click()
                print(filter.text)
                time.sleep(3)
                print('Applied Filters ,success')
                break
            else:
                print('element not found on {}'.format(i))

        # select price
        print('selecting price')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a')))
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a').click()
        time.sleep(3)

        try:
            # set price < 300$
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/span').click()
            time.sleep(3)
            for i in range(0, 10):
                print(i)
                element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/ul/li[' + str(i) + ']'
                print(element)
                try:
                    filter = driver.find_element_by_xpath(element).text
                except:
                    print('not such element exists')
                    filter = 'NO'

                if filter == '300K':  #  300K
                    mainelem = element
                    filter = driver.find_element_by_xpath(mainelem)
                    filter.click()
                    print(filter.text)
                    time.sleep(3)
                    print('Applied Filters ,success')
                    time.sleep(3)
                    print('clicking on botton')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[3]/div/div/a[2]').click()
                    break
                else:
                    print('element not found on {}'.format(i))
        except:
            print('price element not found')

        print('locating basic info about communities')
        time.sleep(3)
        try:
            driver = self.driver
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a')))
            self.dict_lennar_filter_info['Communities num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a').text
            index1 = self.dict_lennar_filter_info['Communities num'].find('(')
            index2 = self.dict_lennar_filter_info['Communities num'].find(')')
            self.dict_lennar_filter_info['Communities num'] = self.dict_lennar_filter_info['Communities num'][index1 + 1:index2]
            self.dict_lennar_filter_info['Quick Move-In Homes num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[2]/a').text
            self.dict_lennar_filter_info['Floorplans num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[3]/a').text
            self.num_of_communities = self.dict_lennar_filter_info['Communities num']
            print('Communities number is {}'.format(self.dict_lennar_filter_info['Communities num']))
            print('change view to list')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a')))
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
            time.sleep(3)
            print('Basic Information scanned {}'.format(self.dict_lennar_filter_info))
        except:
            print('Failed, to Locate information from Lennar')

        try:
            print('copy to xls basic info')
            print('Creating new xls sheet')
            wb = openpyxl.load_workbook(self.xls_name)
            if wb.sheetnames.count(self.metropolitan + ' comm_data') == 0:
                example_sheet = wb['comm_data']
                wb.copy_worksheet(example_sheet)
                new_sheet = wb['comm_data Copy']
                new_sheet.title = self.metropolitan + ' comm_data'
                wb.save(self.xls_name)
                print("XLS new sheet is ready, sheet name: {}".format(new_sheet.title))
                wb.close()
            else:
                print('address was exist in xls')
        except:
            print('failed to connect to xls')

        try:
            time.sleep(3)
            print('opening xls')
            print('xls name {}'.format(self.xls_name))
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.metropolitan + ' comm_data']
            sheet['K2'].value = self.dict_lennar_filter_info['metropolitan name']
            sheet['L2'].value = self.dict_lennar_filter_info['Communities num']
            sheet['M2'].value = self.dict_lennar_filter_info['Quick Move-In Homes num']
            sheet['N2'].value = self.dict_lennar_filter_info['Floorplans num']
            sheet['J2'].value = datetime.datetime.now()
            wb.save(self.xls_name)
            wb.close()
            print('sheet name is {}'.format(self.metropolitan + ' comm_data'))
            print('basic community info bar was saved in xls')
            return True
        except:
            print('failed to copy basic community info to XLS ')
            logging.debug('failed to open XLS')
            return False

    '''
        def community_and_homes_all_data_to_xls_and_SQL(self):
        try:
            print('Calculating the num of Pages to scroll - communities')
            if int(self.num_of_communities) < 30:
                self.num_of_comm_pages = 1
                print('Num of communities {}'.format(self.num_of_communities))
                print('Num of pages of communities {}'.format(self.num_of_comm_pages))
            else:
                if int(self.num_of_communities) < 60:
                    self.num_of_comm_pages = 2
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
                else:
                    self.num_of_comm_pages = int(self.num_of_communities) / 30
                    self.num_of_comm_pages = round(self.num_of_comm_pages)
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
        except:
            print('could not calculate data about communities')

        if int(self.num_of_communities) < 30:  # if communities < 30 (one page)
            print('if communities < 30')
            time.sleep(2)
            for x in range(0, int(self.num_of_communities)):  # int(self.num_of_communities):
                print('community area entered')
                try:
                    driver = self.driver
                    time.sleep(5)
                    print('change view to list')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                    time.sleep(5)
                except:
                    print('list button not located')
                try:
                    driver = self.driver
                    print('Preparing to Enter community on num {}'.format(x))
                    time.sleep(10)
                    print('trying to locate community address')
                    x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[2]'
                    print(x_path)
                    self.addr = driver.find_element_by_xpath(x_path).text
                    self.dict_community_data['address'] = self.addr
                    print('Community Address: {}'.format(self.addr))
                    self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[1]/a/strong'
                    self.name = driver.find_element_by_xpath(self.x_path_name).text
                    self.dict_community_data['name_community'] = self.name
                    print('Community Name: {}'.format(self.name))
                    self.community_address_list_full.append(self.dict_community_data['address'])
                    print('Community address was added to list for automation')
                    print('scrolling')
                    scroll = 245 * x
                    print(scroll)
                    scroll = "window.scrollTo(0, " + str(scroll) + ")"
                    driver.execute_script(scroll)
                    time.sleep(10)
                    print('scrolled'.format(x))
                    print('trying to click the scrolled community')
                    print(driver.current_url)
                    print(self.x_path_name)
                    print('clicking')
                    driver.find_element_by_xpath(self.x_path_name).click()
                    time.sleep(3)
                    print('clicked')
                    time.sleep(10)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                    print('SUCCESS - community found and pressed')
                    time.sleep(10)
                    c = 0
                except:
                    print('FAILED - to locate community on xpath num {}'.format(x))
                    c = 1

                # if community is located
                if c == 0:
                    print('After Community was located - starting to download data')
                    print('1 First generating ID for Community')
                    time.sleep(5)
                    self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                    self.id_random_list.append(self.dict_community_data['id_generated'])
                    print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                    try:
                        print('2 try copy overview data')
                        driver = self.driver
                        self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                    except:
                        print('failed to locate overview')

                    time.sleep(5)
                    try:
                        print('3 try copy picture 1 map ')
                        driver = self.driver
                        self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                        print('actually downloading the image and changing the name.jpg')
                        urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                    except:
                        print('failed to locate pictures map')
                        self.dict_community_data['community_map_url'] = 'NA'
                    try:
                        print('4 try copy pictures 2')
                        driver = self.driver
                        self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                        urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                    except:
                        print('failed to locate pictures 2')
                        self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                    try:
                        print('5 Available Homes and floorplans')
                        driver = self.driver
                        self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                        self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                        print('success to copy home toolbar data {}'.format(self.dict_community_data))
                    except:
                        print('failed to locate homes toolbar')

                    try:
                        print('6 copy community data to xls')
                        print('open xls '.format(self.xls_name))
                        wb = openpyxl.load_workbook(self.xls_name)
                        time.sleep(2)
                        sheet = wb[self.metropolitan + ' comm_data']
                        sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                        sheet['B' + str(self.row)].value = self.dict_community_data['address']
                        sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                        sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                        sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                        sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                        sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                        sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                        sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                        wb.save(self.xls_name)
                        wb.close()
                        print('COMMUNITY DATA - saved in xls')
                        self.row = self.row + 1
                    except:
                        print('failed to copy community data to XLS ')

                    # homes general data
                    try:
                        driver = self.driver
                        print('scrolling to homes')
                        driver.execute_script("window.scrollTo(0, 2050)")
                        time.sleep(5)
                        print('changing view to list')
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                        print('Calculating num of homes')
                        self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                        print('num of homes to verify {}'.format(self.num_of_moving_homes))
                        print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans'][12:-1]))
                        self.floorplan_homes = self.dict_community_data['available_homes_floorplans'][12:-1]
                        time.sleep(3)
                    except:
                        print('could not locate general homes information')

                    print('copy homes + floorpans :):):):):):)')
                    print('FLOORPLANS')
                    for j in range(0, int(self.floorplan_homes)):
                        try:
                            driver = self.driver
                            print('Choosing floorplans Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                            print('floorplans clicked')
                            time.sleep(5)
                            print('For floorplan - Home number {}'.format(j + 1))
                            time.sleep(3)
                            print('Scrolling to Home')
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to floorplans Home')
                            time.sleep(3)
                            self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                        except:
                            print('could not locate floorplan home!')

                        print('floorplans - trying to enter - Homes')

                        if int(self.floorplan_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if floorplans home is <= 1 , trying to find home link')
                                print('clicking on floorplans home link')
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans home entered')
                                print('waiting for the floorplans home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans Home LOCATED in the list!')
                                time.sleep(15)
                                print('first generating ID floorplans for home')
                                time.sleep(5)
                                print('generating home floorplans id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans could not locate home link <= 1')
                        else:
                            try:
                                driver = self.driver
                                print('IF floorplans Homes count more than > 1')
                                print('clicking on floorplans home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home floorplan LOCATED!')
                                time.sleep(15)
                                print('generating floorplans home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans Home not located on path number {}'.format(j + 1))

                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "TBB"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                            time.sleep(2)
                        except:
                            print('failed to generate')

                        try:
                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                self.dict_home_data['availability'] = 'NA'
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')

                            # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                        except:
                            print('could not locate HOME / elements')

                        print('Trying to copy all gained Homes data to XLS file')
                        try:
                            print('xls - creating new sheet with home name')
                            wb = openpyxl.load_workbook(self.xls_name)
                            if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                                print('creating xls')
                                example_sheet = wb['home_data']
                                wb.copy_worksheet(example_sheet)
                                print(wb.sheetnames)
                                new_sheet = wb['home_data Copy']
                                new_sheet.title = self.metropolitan + ' home_data'
                                wb.save(self.xls_name)
                                print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                                print(wb.sheetnames)
                                wb.close()
                            else:
                                print("Metropolitan Homes sheet already created in xls")
                        except:
                            print('failed to connect to xls file and create sheet')

                        # copy home basic info to xls
                        try:
                            # opening xls
                            print('IMPORTANT - copy home info to xls')
                            wb = openpyxl.load_workbook(self.xls_name)
                            sheet = wb[self.metropolitan + ' home_data']
                            sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                            sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                            sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                            sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                            sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                            sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                            sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                            sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                            sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                            sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                            sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                            sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                            sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                            sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                            sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                            sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                            sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                            sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                            sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']

                            wb.save(self.xls_name)
                            wb.close()
                            print('xls floorplan - HOME params was saved')
                            self.rowhome = self.rowhome + 1
                        except:
                            print('failed to copy floorplans HOME params to xls')
                            logging.debug('failed to open XLS')

                        print('Trying to Connect and copy same data to MySQL server')
                        self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                        try:
                            db = mysql.connector.connect(
                                host='107.180.21.18',
                                user='grow097365',
                                passwd='Jknm678##Tg',
                                database='equity_property'
                            )
                            mycursor = db.cursor()
                            print(db)  # checking our connection to DB
                            command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                            print(command)

                            mycursor.execute(command)
                            myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                            print(len(myresult))
                            print(myresult)

                            if len(myresult) == 0:
                                print('Similar homes not found, copying to database!')
                                db = mysql.connector.connect(
                                    host='107.180.21.18',
                                    user='grow097365',
                                    passwd='Jknm678##Tg',
                                    database='equity_property'
                                )
                                mycursor = db.cursor()
                                print(db)
                                sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                val = (self.dict_home_data['id_generated'],
                                       datetime.datetime.now(),
                                       self.dict_home_data['address'],
                                       self.short_state,
                                       self.metropolitan,
                                       self.dict_home_data['home_name'],
                                       self.dict_home_data['home_size'],
                                       self.dict_home_data['beds'],
                                       self.dict_home_data['baths'],
                                       self.dict_home_data['garage'],
                                       self.dict_home_data['priced_from'],
                                       self.dict_home_data['gallery_view_picture'],
                                       self.dict_home_data['type'],
                                       str(self.dict_home_data['id_generated_home']),
                                       self.dict_home_data['name_community'])
                                mycursor.execute(sql, val)
                                db.commit()
                                time.sleep(3)
                                print('IMPORTANT - Home floorplan data copied to mySQL')
                            else:
                                print('Similar home found in database')
                        except:
                            print('failed to work with mySQL')

                        try:
                            driver = self.driver
                            print('trying to go back to HOMES list after data copied')
                            driver.back()
                            time.sleep(7)
                        except:
                            print('could not go back on general HOMES list')
                    print('HOMES')
                    for j in range(0, int(self.num_of_moving_homes)):
                        try:
                            driver = self.driver
                            print('entering Homes and copy the data')
                            print('For Home number {}'.format(j + 1))
                            print('Choosing quick mov in Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                            time.sleep(3)
                            print('Scrolling to Home')
                            time.sleep(3)
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to Homes')
                            print('trying to enter - Homes')
                        except:
                            print('could not locate floorplan home!')

                        if int(self.num_of_moving_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if home is <= 1 , trying to find home link')
                                print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('home link clicked')
                                time.sleep(5)
                                print('home entered')
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('could not locate home')
                        else:
                            try:
                                driver = self.driver
                                print('if Homes more then > 1')
                                print('clicking on home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                time.sleep(5)
                                print('home link clicked')
                                print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('Home not located on path number {}'.format(j + 1))
                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "MIR"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))

                            time.sleep(2)
                            try:
                                print('try copy home picture')
                                driver = self.driver
                                print('getting the source link of the picture')
                                self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                                print(self.dict_home_data['gallery_view_picture'])
                                urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                            except:
                                print('failed to locate pictures')
                                self.dict_home_data['gallery_view_picture'] = 'NA'

                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                driver = self.driver
                                self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                                print(self.dict_home_data['baths'])
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')
                        except:
                            print('could not locate HOME / elements')
                    print('after all homes was scanned, we going back to community')
                    try:
                        driver = self.driver
                        time.sleep(5)
                        driver.back()
                        time.sleep(5)
                        driver.back()
                        time.sleep(10)
                        print('Waiting till the page will load the community')
                    except:
                        print('could not go back on community list')
            print('END of work on communities < 30')
        else:
            print('if communities > 30 and we got pages to scroll')
            time.sleep(2)
            for page in range(self.num_of_comm_pages):
                print('Comm page num {}'.format(page + 1))
                for self.row_num_xls in range(0, 29):  # 30 communities per page
                    print('community area entered')
                    try:
                        driver = self.driver
                        print('change view to list')
                        time.sleep(5)
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                    except:
                        print('list button not located')
                    try:
                        driver = self.driver
                        print('Preparing to Enter community on num {}'.format(self.row_num_xls))
                        time.sleep(10)
                        print('trying to locate community address')
                        x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[2]'
                        print(x_path)
                        self.addr = driver.find_element_by_xpath(x_path).text
                        self.dict_community_data['address'] = self.addr
                        print('Community Address: {}'.format(self.addr))
                        self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[1]/a/strong'
                        self.name = driver.find_element_by_xpath(self.x_path_name).text
                        self.dict_community_data['name_community'] = self.name
                        print('Community Name: {}'.format(self.name))
                        self.community_address_list_full.append(self.dict_community_data['address'])
                        print('Community address was added to list for automation')

                        print('scrolling')
                        scroll = 245 * self.row_num_xls
                        print(scroll)
                        scroll = "window.scrollTo(0, " + str(scroll) + ")"
                        driver.execute_script(scroll)
                        time.sleep(10)
                        print('scrolled'.format(self.row_num_xls))

                        print('trying to click the scrolled community')
                        print(driver.current_url)
                        print(self.x_path_name)
                        print('clicking')
                        driver.find_element_by_xpath(self.x_path_name).click()
                        time.sleep(3)
                        print('clicked')
                        time.sleep(10)
                        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        print('SUCCESS - community found and pressed')
                        time.sleep(10)
                        c = 0
                    except:
                        print('FAILED - to locate community on xpath num {}'.format(self.row_num_xls))
                        c = 1

                    # if community is located
                    if c == 0:
                        print('After Community was located - starting to download data')
                        print('1 First generating ID for Community')
                        time.sleep(5)
                        self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                        self.id_random_list.append(self.dict_community_data['id_generated'])
                        print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                        try:
                            print('2 try copy overview data')
                            driver = self.driver
                            self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                        except:
                            print('failed to locate overview')

                        time.sleep(5)
                        try:
                            print('3 try copy picture 1 map ')
                            driver = self.driver
                            self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                            print('actually downloading the image and changing the name.jpg')
                            urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                        except:
                            print('failed to locate pictures map')
                            self.dict_community_data['community_map_url'] = 'NA'
                        try:
                            print('4 try copy pictures 2')
                            driver = self.driver
                            self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                        except:
                            print('failed to locate pictures 2')
                            self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                        try:
                            print('5 Available Homes and floorplans')
                            driver = self.driver
                            self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                            self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                                                                                                                 # //*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]
                            print('success to copy home toolbar data {}'.format(self.dict_community_data))
                        except:
                            print('failed to locate homes toolbar')

                        try:
                            print('6 copy community data to xls')
                            print('open xls '.format(self.xls_name))
                            wb = openpyxl.load_workbook(self.xls_name)
                            time.sleep(2)
                            sheet = wb[self.metropolitan + ' comm_data']
                            sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                            sheet['B' + str(self.row)].value = self.dict_community_data['address']
                            sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                            sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                            sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                            sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                            sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                            sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                            sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                            wb.save(self.xls_name)
                            wb.close()
                            print('COMMUNITY DATA - saved in xls')
                            self.row = self.row + 1
                        except:
                            print('failed to copy community data to XLS ')

                        try:
                            print('# homes general data')
                            driver = self.driver
                            print('scrolling to homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(5)
                            print('changing view to list')
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                            time.sleep(5)
                            print('Calculating num of homes')
                            self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                            self.floorplan_homes = self.dict_community_data['available_homes_floorplans'][-2:-1]
                            print('num of homes to verify {}'.format(self.num_of_moving_homes))
                            print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans']))
                            time.sleep(3)
                        except:
                            print('could not locate general homes information')

                        print('copy homes + floorpans :):):):):):)')
                        print('FLOORPLANS')
                        for j in range(0, int(self.floorplan_homes)):
                            try:
                                driver = self.driver
                                print('Choosing floorplans Homes')
                                driver.execute_script("window.scrollTo(0, 2050)")
                                time.sleep(3)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                                print('floorplans clicked')
                                time.sleep(5)
                                print('For floorplan - Home number {}'.format(j + 1))
                                time.sleep(3)
                                print('Scrolling to Home')
                                scroll = 2000 + (245 * j)
                                scroll = "window.scrollTo(0, " + str(scroll) + ")"
                                driver.execute_script(scroll)
                                print('scrolled to floorplans Home')
                                time.sleep(3)
                                self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                                urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                            except:
                                print('could not locate floorplan home!')

                            print('floorplans - trying to enter - Homes')

                            if int(self.floorplan_homes) <= 1:
                                try:
                                    driver = self.driver
                                    print('if floorplans home is <= 1 , trying to find home link')
                                    print('clicking on floorplans home link')
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('floorplans home link clicked')
                                    time.sleep(5)
                                    print('floorplans home entered')
                                    print('waiting for the floorplans home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('floorplans Home LOCATED in the list!')
                                    time.sleep(15)
                                    print('first generating ID floorplans for home')
                                    time.sleep(5)
                                    print('generating home floorplans id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('floorplans could not locate home link <= 1')
                            else:
                                try:
                                    driver = self.driver
                                    print('IF floorplans Homes count more than > 1')
                                    print('clicking on floorplans home link')
                                    ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                    time.sleep(5)
                                    print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                    print('floorplans home link clicked')
                                    time.sleep(5)
                                    print('floorplans waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home floorplan LOCATED!')
                                    time.sleep(15)
                                    print('generating floorplans home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('floorplans Home not located on path number {}'.format(j + 1))

                            try:
                                self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                                self.dict_home_data['type'] = "TBB"
                                print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                                time.sleep(2)
                            except:
                                print('failed to generate')

                            try:
                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                    print(self.dict_home_data['home_name'])
                                except:
                                    print('home name not found')

                                try:
                                    self.dict_home_data['address'] = self.dict_community_data['address']
                                    print(self.dict_home_data['address'])
                                except:
                                    print('address not found')

                                try:
                                    self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                    print(self.dict_home_data['name_community'])
                                except:
                                    print('name community not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['home_site'])
                                except:
                                    print('home site not found')

                                self.dict_home_data['included_features_pdf'] = 'under solution'

                                try:
                                    self.dict_home_data['availability'] = 'NA'
                                    print(self.dict_home_data['availability'])
                                except:
                                    print('availability not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                    self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                    print(self.dict_home_data['priced_from'])
                                except:
                                    print('priced from not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['home_size'])
                                except:
                                    print('home size not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                    print(self.dict_home_data['stories'])
                                except:
                                    print('stories not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                    print(self.dict_home_data['beds'])
                                except:
                                    print('beds not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                    print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                                except:
                                    print('baths not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                    print(self.dict_home_data['garage'])
                                except:
                                    print('garage not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                    print(self.dict_home_data['description'])
                                except:
                                    print('description not found')

                                try:
                                    driver = self.driver
                                    print('trying to copy home FloorPlan Pic scrolling')
                                    driver.execute_script("window.scrollTo(0, 1600)")
                                    time.sleep(4)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                    time.sleep(3)
                                    self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                                except:
                                    print('could not locate home pics and FloorPlan Pic')

                                # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                            except:
                                print('could not locate HOME / elements')

                            print('Trying to copy all gained Homes data to XLS file')
                            try:
                                print('xls - creating new sheet with home name')
                                wb = openpyxl.load_workbook(self.xls_name)
                                if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                                    print('creating xls')
                                    example_sheet = wb['home_data']
                                    wb.copy_worksheet(example_sheet)
                                    print(wb.sheetnames)
                                    new_sheet = wb['home_data Copy']
                                    new_sheet.title = self.metropolitan + ' home_data'
                                    wb.save(self.xls_name)
                                    print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                                    print(wb.sheetnames)
                                    wb.close()
                                else:
                                    print("Metropolitan Homes sheet already created in xls")
                            except:
                                print('failed to connect to xls file and create sheet')

                            # copy home basic info to xls
                            try:
                                # opening xls
                                print('IMPORTANT - copy home info to xls')
                                wb = openpyxl.load_workbook(self.xls_name)
                                sheet = wb[self.metropolitan + ' home_data']
                                sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                                sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                                sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                                sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                                sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                                sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                                sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                                sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                                sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                                sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                                sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                                sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                                sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                                sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                                sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                                sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                                sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                                sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                                sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']

                                wb.save(self.xls_name)
                                wb.close()
                                print('xls floorplan - HOME params was saved')
                                self.rowhome = self.rowhome + 1
                            except:
                                print('failed to copy floorplans HOME params to xls')
                                logging.debug('failed to open XLS')

                            print('Trying to Connect and copy same data to MySQL server')
                            self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                            try:
                                db = mysql.connector.connect(
                                    host='107.180.21.18',
                                    user='grow097365',
                                    passwd='Jknm678##Tg',
                                    database='equity_property'
                                )
                                mycursor = db.cursor()
                                print(db)  # checking our connection to DB
                                command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                                print(command)

                                mycursor.execute(command)
                                myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                                print(len(myresult))
                                print(myresult)

                                if len(myresult) == 0:
                                    print('Similar homes not found, copying to database!')
                                    db = mysql.connector.connect(
                                        host='107.180.21.18',
                                        user='grow097365',
                                        passwd='Jknm678##Tg',
                                        database='equity_property'
                                    )
                                    mycursor = db.cursor()
                                    print(db)
                                    sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                    val = (self.dict_home_data['id_generated'],
                                           datetime.datetime.now(),
                                           self.dict_home_data['address'],
                                           self.short_state,
                                           self.metropolitan,
                                           self.dict_home_data['home_name'],
                                           self.dict_home_data['home_size'],
                                           self.dict_home_data['beds'],
                                           self.dict_home_data['baths'],
                                           self.dict_home_data['garage'],
                                           self.dict_home_data['priced_from'],
                                           self.dict_home_data['gallery_view_picture'],
                                           self.dict_home_data['type'],
                                           str(self.dict_home_data['id_generated_home']),
                                           self.dict_home_data['name_community'])
                                    mycursor.execute(sql, val)
                                    db.commit()
                                    time.sleep(3)
                                    print('IMPORTANT - Home floorplan data copied to mySQL')
                                else:
                                    print('Similar home found in database')
                            except:
                                print('failed to work with mySQL')

                            try:
                                driver = self.driver
                                print('trying to go back to HOMES list after data copied')
                                driver.back()
                                time.sleep(7)
                            except:
                                print('could not go back on general HOMES list')
                        print('HOMES')
                        for j in range(0, int(self.num_of_moving_homes)):
                            try:
                                driver = self.driver
                                print('entering Homes and copy the data')
                                print('For Home number {}'.format(j + 1))
                                print('Choosing quick mov in Homes')
                                driver.execute_script("window.scrollTo(0, 2050)")
                                time.sleep(3)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                                time.sleep(3)
                                print('Scrolling to Home')
                                time.sleep(3)
                                scroll = 2000 + (245 * j)
                                scroll = "window.scrollTo(0, " + str(scroll) + ")"
                                driver.execute_script(scroll)
                                print('scrolled to Homes')
                                print('trying to enter - Homes')
                            except:
                                print('could not locate floorplan home!')

                            if int(self.num_of_moving_homes) <= 1:
                                try:
                                    driver = self.driver
                                    print('if home is <= 1 , trying to find home link')
                                    print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('home link clicked')
                                    time.sleep(5)
                                    print('home entered')
                                    print('waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home LOCATED!')
                                    time.sleep(15)
                                    print('generating home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('could not locate home')
                            else:
                                try:
                                    driver = self.driver
                                    print('if Homes more then > 1')
                                    print('clicking on home link')
                                    ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                    time.sleep(5)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                    time.sleep(5)
                                    print('home link clicked')
                                    print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                    print('waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home LOCATED!')
                                    time.sleep(15)
                                    print('generating home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('Home not located on path number {}'.format(j + 1))
                            try:
                                self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                                self.dict_home_data['type'] = "MIR"
                                print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))

                                time.sleep(2)
                                try:
                                    print('try copy home picture')
                                    driver = self.driver
                                    print('getting the source link of the picture')
                                    self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                                    print(self.dict_home_data['gallery_view_picture'])
                                    urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                                except:
                                    print('failed to locate pictures')
                                    self.dict_home_data['gallery_view_picture'] = 'NA'

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                    print(self.dict_home_data['home_name'])
                                except:
                                    print('home name not found')

                                try:
                                    self.dict_home_data['address'] = self.dict_community_data['address']
                                    print(self.dict_home_data['address'])
                                except:
                                    print('address not found')

                                try:
                                    self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                    print(self.dict_home_data['name_community'])
                                except:
                                    print('name community not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                    print(self.dict_home_data['home_site'])
                                except:
                                    print('home site not found')

                                self.dict_home_data['included_features_pdf'] = 'under solution'

                                try:
                                    driver = self.driver
                                    self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['availability'])
                                except:
                                    print('availability not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                    self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                    print(self.dict_home_data['priced_from'])
                                except:
                                    print('priced from not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                    print(self.dict_home_data['home_size'])
                                except:
                                    print('home size not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                    print(self.dict_home_data['stories'])
                                except:
                                    print('stories not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                    print(self.dict_home_data['beds'])
                                except:
                                    print('beds not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                                    print(self.dict_home_data['baths'])
                                except:
                                    print('baths not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                                    print(self.dict_home_data['garage'])
                                except:
                                    print('garage not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                    print(self.dict_home_data['description'])
                                except:
                                    print('description not found')

                                try:
                                    driver = self.driver
                                    print('trying to copy home FloorPlan Pic scrolling')
                                    driver.execute_script("window.scrollTo(0, 1600)")
                                    time.sleep(4)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                    time.sleep(3)
                                    self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                                except:
                                    print('could not locate home pics and FloorPlan Pic')
                            except:
                                print('could not locate HOME / elements')

                        print('after all homes was scanned, we going back to community')
                        try:
                            driver = self.driver
                            time.sleep(5)
                            driver.back()
                            time.sleep(5)
                            driver.back()
                            time.sleep(10)
                            print('Waiting till the page will load the community')
                        except:
                            print('could not go back on community list')
                    else:
                        print('End of community list, number of communities was {} '.format(self.num_of_communities))
                # page num
                try:
                    driver = self.driver
                    print('scrolling to next page button')
                    driver.execute_script("window.scrollTo(0, 8000)")
                    time.sleep(5)

                    for i in range(10, 0, -1):
                        try:
                            print('trying to locate next page button {}'.format(i))
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[2]/a[' + str(i) + ']').click()
                            time.sleep(5)
                            print('next button located on number {}'.format(i))
                        except:
                            print('trying another path - next page button was not located')

                    print('next page pressed')
                    time.sleep(2)
                    driver.execute_script("window.scrollTo(0, 0)")
                    print('scrolling back to top')
                    time.sleep(6)
                except:
                    print('button not located')
            print('END of work on communities > 30')
    '''

    def return_community_address_list(self):
        return self.community_address_list_full

    def return_Generated_Id_list(self):
        return self.id_random_list

    def return_list_of_homes(self):
        return self.list_of_homes
class CMA(object):  # Zillow API - api_key = "X1-ZWz1hbswvtw74b_3tnpx"
    def __init__(self, address, api_key, zip_code, xls_name):
        # all setup params
        self.address = address
        self.api_key = api_key
        self.zip_code = zip_code
        self.xls_name = xls_name

        self.dict_zillow = {
            'address': self.address,
            'api_key': self.api_key,
            'zip_code': self.zip_code,
            'zpid': '',
            'link - comparables': '',
            'link - graphs_and_data': '',
            'link - home_details': '',
            'link - map_this_home': '',
            'amount': '',
            'amount_currency': '',
            'amount_last_updated': '',
            'valuation_range_high': '',
            'valuation_range_low': '',
            'bathrooms': '',
            'bedrooms': '',
            'complete': '',
            'finished_sqft': '',
            'fips_county': '',
            'last_sold_date': '',
            'last_sold_price': '',
            'lot_size_sqft': '',
            'tax_assessment': '',
            'tax_assessment_year': '',
            'usecode': '',
            'year_built': '',

        }

    # activation zillow API and copy params to dict
    def zillow_api(self):
        try:
            locale.setlocale(locale.LC_ALL, '')
            # api = zillow.ValuationApi()
            # get deep search results, also getting the zswid-ID
            # data = api.GetDeepSearchResults(self.api_key, self.address, self.zip_code)
            # my_dict = data.get_dict()
            # copy web page params to dictionary dict_zillow
            # self.dict_zillow['zpid'] = my_dict['zpid']
            # self.dict_zillow['link - comparables'] = my_dict['links']['comparables']
            # self.dict_zillow['link - graphs_and_data'] = my_dict['links']['graphs_and_data']
            # self.dict_zillow['link - home_details'] = my_dict['links']['home_details']
            # self.dict_zillow['link - map_this_home'] = my_dict['links']['map_this_home']
            # self.dict_zillow['amount'] = my_dict['zestimate']['amount']
            # self.dict_zillow['amount_currency'] = my_dict['zestimate']['amount_currency']
            # self.dict_zillow['amount_last_updated'] = my_dict['zestimate']['amount_last_updated']
            # self.dict_zillow['valuation_range_high'] = my_dict['zestimate']['valuation_range_high']
            # self.dict_zillow['valuation_range_low'] = my_dict['zestimate']['valuation_range_low']
            # self.dict_zillow['bathrooms'] = my_dict['extended_data']['bathrooms']
            # self.dict_zillow['bedrooms'] = my_dict['extended_data']['bedrooms']
            # self.dict_zillow['complete'] = my_dict['extended_data']['complete']
            # self.dict_zillow['finished_sqft'] = my_dict['extended_data']['finished_sqft']
            # self.dict_zillow['fips_county'] = my_dict['extended_data']['fips_county']
            # self.dict_zillow['last_sold_date'] = my_dict['extended_data']['last_sold_date']
            # self.dict_zillow['last_sold_price'] = my_dict['extended_data']['last_sold_price']
            # self.dict_zillow['lot_size_sqft'] = my_dict['extended_data']['lot_size_sqft']
            # self.dict_zillow['tax_assessment_year'] = my_dict['extended_data']['tax_assessment_year']
            # self.dict_zillow['usecode'] = my_dict['extended_data']['usecode']
            # self.dict_zillow['year_built'] = my_dict['extended_data']['year_built']
            return True
        except:
            print('fail to get params from zillow api')
            logging.debug('fail')
            self.dict_zillow['zpid'] = 'fail to get params from zillow api'
            return False

    # printing all dicts
    def print_all(self):
        pp = pprint.PrettyPrinter(indent=4)
        pp.pprint(self.dict_zillow)

    # returning all dictionaries for future use for general list
    def return_dict_zillow(self):
        return self.dict_zillow

    # copy dict to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.address[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.address[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.address[:25]]
            # print(wb.sheetnames)
            sheet['B2'].value = self.dict_zillow['address']
            sheet['B3'].value = self.dict_zillow['api_key']
            sheet['B4'].value = self.dict_zillow['zip_code']
            sheet['B5'].value = self.dict_zillow['zpid']
            sheet['B7'].value = self.dict_zillow['link - comparables']
            sheet['B8'].value = self.dict_zillow['link - graphs_and_data']
            sheet['B9'].value = self.dict_zillow['link - home_details']
            sheet['B10'].value = self.dict_zillow['link - map_this_home']
            sheet['B12'].value = self.dict_zillow['amount']
            sheet['B13'].value = self.dict_zillow['amount_currency']
            sheet['B14'].value = self.dict_zillow['amount_last_updated']
            sheet['B15'].value = self.dict_zillow['valuation_range_high']
            sheet['B16'].value = self.dict_zillow['valuation_range_low']
            sheet['B18'].value = self.dict_zillow['bathrooms']
            sheet['B19'].value = self.dict_zillow['bedrooms']
            sheet['B20'].value = self.dict_zillow['complete']
            sheet['B21'].value = self.dict_zillow['finished_sqft']
            sheet['B22'].value = self.dict_zillow['fips_county']
            sheet['B23'].value = self.dict_zillow['last_sold_date']
            sheet['B24'].value = self.dict_zillow['last_sold_price']
            sheet['B25'].value = self.dict_zillow['lot_size_sqft']
            sheet['B26'].value = self.dict_zillow['tax_assessment']
            sheet['B27'].value = self.dict_zillow['tax_assessment_year']
            sheet['B28'].value = self.dict_zillow['usecode']
            sheet['B29'].value = self.dict_zillow['year_built']
            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Dictionaries was completed & saved in {}".format(self.xls_name))
            logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
            return True
        except:
            return False
























